from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from review_analytics.cli import AppDependencies, run
from review_analytics.config import AppConfig
from review_analytics.repositories import AnalysisRepository, ReviewRepository, initialize_database
from tests.fakes import FakeGeminiClient


def test_full_offline_pipeline_covers_every_command_and_output_without_network(tmp_path, monkeypatch, capsys):
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)
    database_path = tmp_path / "data" / "reviews.db"
    database_path.parent.mkdir(parents=True)
    initialize_database(database_path)
    reviews = ReviewRepository(database_path)
    analyses = AnalysisRepository(database_path)
    config = AppConfig(
        database_path=database_path,
        analysis_batch_size=7,
        extraction_chunk_characters=500,
        ai_retry_count=0,
        chart_font_candidates=("DejaVu Sans",),
        log_file=tmp_path / "logs" / "app.log",
        output_directory=tmp_path / "output",
    )
    clients: list[FakeGeminiClient] = []

    def client_factory() -> FakeGeminiClient:
        client = FakeGeminiClient()
        clients.append(client)
        return client

    dependencies = AppDependencies(config, reviews, analyses, client_factory)
    sample_csv = Path(__file__).parents[1] / "fixtures" / "sample_reviews.csv"
    output_dir = tmp_path / "dashboard"
    csv_output = tmp_path / "exports" / "reviews.csv"
    xlsx_output = tmp_path / "exports" / "reviews.xlsx"

    assert run(["import", "--file", str(sample_csv)], dependencies) == 0
    assert run(["clean", "--pending"], dependencies) == 0
    assert run(["analyze", "--unanalyzed"], dependencies) == 0
    assert run(["extract"], dependencies) == 0
    assert run(["list", "--page", "1", "--size", "5"], dependencies) == 0
    assert run(["show", "1"], dependencies) == 0
    assert run(["stats"], dependencies) == 0
    assert run(["dashboard", "--output-dir", str(output_dir)], dependencies) == 0
    assert run(["export", "--format", "csv", "--output", str(csv_output)], dependencies) == 0
    assert run(["export", "--format", "xlsx", "--output", str(xlsx_output)], dependencies) == 0

    console = capsys.readouterr().out
    assert "스킵: 2건" in console
    assert "거절: 4건" in console
    assert "positive" in console or "negative" in console or "neutral" in console
    assert "고객 리뷰 감정 분석 종합 리포트" in console

    with sqlite3.connect(database_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM raw_reviews").fetchone()[0] == 32
        assert connection.execute("SELECT COUNT(*) FROM clean_reviews").fetchone()[0] == 28
        assert connection.execute("SELECT COUNT(*) FROM sentiment_analyses").fetchone()[0] == 28
        assert connection.execute("SELECT COUNT(*) FROM insight_extractions").fetchone()[0] == 1

    expected_dashboard_files = {
        "sentiment_distribution.png",
        "sentiment_trend.png",
        "rating_sentiment_matrix.png",
        "review_sentiment_report.md",
    }
    assert {path.name for path in output_dir.iterdir()} == expected_dashboard_files
    assert all((output_dir / name).stat().st_size > 0 for name in expected_dashboard_files)
    report = (output_dir / "review_sentiment_report.md").read_text(encoding="utf-8")
    assert "오프라인 Fake Gemini 요약" in report
    assert "근거 리뷰" in report

    assert csv_output.read_bytes().startswith(b"\xef\xbb\xbf")
    assert len(csv_output.read_text(encoding="utf-8-sig").splitlines()) == 29
    workbook = load_workbook(xlsx_output, read_only=True)
    try:
        assert workbook.active.max_row == 29
    finally:
        workbook.close()

    assert len(clients) == 2
    assert all(client.closed for client in clients)
    log_text = config.log_file.read_text(encoding="utf-8")
    assert "GEMINI_API_KEY" not in log_text
    assert "오프라인 Fake Gemini 요약" not in log_text

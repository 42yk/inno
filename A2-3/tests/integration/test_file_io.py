from pathlib import Path
import gc
import logging
import sqlite3
import warnings
import zipfile
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook

from review_analytics.dto import ExportRow
from review_analytics.errors import InputFileError, OutputWriteError
from review_analytics.models import ExportFormat, Sentiment

from .conftest import scalar


def test_read_reviews_maps_csv_rows_to_raw_inputs_without_cleaning(tmp_path):
    """Normalizing source cells while reading would lose the raw import record."""
    from review_analytics.file_io.reader import read_reviews

    path = tmp_path / "reviews.csv"
    path.write_text(
        "review_text,rating,review_date,product_name\n  Great   item  ,5,2026-08-01,Bottle\n",
        encoding="utf-8",
    )

    rows = read_reviews(path)

    assert len(rows) == 1
    assert rows[0].review_text_raw == "  Great   item  "
    assert rows[0].rating_raw == "5"
    assert rows[0].review_date_raw == "2026-08-01"
    assert rows[0].product_name_raw == "Bottle"
    assert (rows[0].source_type, rows[0].source_ref, rows[0].source_row) == ("csv", "reviews.csv", 2)


def test_read_reviews_maps_xlsx_rows_to_raw_inputs(tmp_path):
    """Returning workbook rows would leak an external object across the File I/O boundary."""
    from review_analytics.file_io.reader import read_reviews

    path = tmp_path / "reviews.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["review_text", "rating", "review_date", "product_name"])
    sheet.append(["좋은   제품", 5, "2026-08-01", "물병"])
    workbook.save(path)

    rows = read_reviews(path)

    assert len(rows) == 1
    assert rows[0].review_text_raw == "좋은   제품"
    assert rows[0].rating_raw == 5
    assert rows[0].source_type == "xlsx"
    assert rows[0].source_row == 2


def test_read_reviews_rejects_missing_required_header_before_returning_rows(tmp_path):
    """Accepting a file without review_text would let ingestion partially persist an invalid file."""
    from review_analytics.file_io.reader import read_reviews

    path = tmp_path / "missing-header.csv"
    path.write_text("rating\n5\n", encoding="utf-8")

    with pytest.raises(InputFileError) as raised:
        read_reviews(path)

    assert raised.value.code == "MISSING_REVIEW_TEXT_COLUMN"


def test_read_reviews_rejects_unsupported_extension(tmp_path):
    """Treating arbitrary text as a review file bypasses the documented input contract."""
    from review_analytics.file_io.reader import read_reviews

    path = tmp_path / "reviews.json"
    path.write_text("[]", encoding="utf-8")

    with pytest.raises(InputFileError) as raised:
        read_reviews(path)

    assert raised.value.code == "UNSUPPORTED_INPUT_FILE_TYPE"


def test_read_reviews_translates_unreadable_xlsx_to_safe_input_error(tmp_path):
    """Leaking a workbook parser exception would expose an external implementation boundary."""
    from review_analytics.file_io.reader import read_reviews

    path = tmp_path / "broken.xlsx"
    path.write_text("not an xlsx workbook", encoding="utf-8")

    with pytest.raises(InputFileError) as raised:
        read_reviews(path)

    assert raised.value.code == "INPUT_FILE_READ_FAILED"


def test_read_reviews_rejects_malformed_quoted_csv_at_the_file_boundary(tmp_path):
    """Permissive CSV parsing would turn a structurally invalid import into a raw review."""
    from review_analytics.file_io.reader import read_reviews

    path = tmp_path / "broken.csv"
    path.write_text('review_text,rating\n"unterminated,5\n', encoding="utf-8")

    with pytest.raises(InputFileError) as raised:
        read_reviews(path)

    assert raised.value.code == "INPUT_FILE_READ_FAILED"


def test_read_reviews_translates_malformed_xlsx_xml_to_safe_input_error(tmp_path):
    """Leaking XML parser details from a workbook would violate the File I/O error boundary."""
    from review_analytics.file_io.reader import read_reviews

    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    workbook.active.append(["review_text"])
    workbook.active.append(["valid row"])
    workbook.save(source)
    path = tmp_path / "broken-xml.xlsx"
    with zipfile.ZipFile(source) as source_archive, zipfile.ZipFile(path, "w") as broken_archive:
        for info in source_archive.infolist():
            content = b"<worksheet>" if info.filename == "xl/worksheets/sheet1.xml" else source_archive.read(info.filename)
            broken_archive.writestr(info, content)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always", ResourceWarning)
        with pytest.raises(InputFileError) as raised:
            read_reviews(path)
        error_code = raised.value.code
        del raised
        gc.collect()

    assert error_code == "INPUT_FILE_READ_FAILED"
    assert not [item for item in caught if item.category is ResourceWarning]


def test_read_reviews_does_not_mask_internal_xlsx_conversion_errors(tmp_path, monkeypatch):
    """Catching local ValueError defects would hide programmer bugs as bad user files."""
    from review_analytics.file_io import reader

    path = tmp_path / "reviews.xlsx"
    workbook = Workbook()
    workbook.active.append(["review_text"])
    workbook.active.append(["valid row"])
    workbook.save(path)

    def fail_conversion(values, index):
        raise ValueError("internal conversion bug")

    monkeypatch.setattr(reader, "_cell_value", fail_conversion)

    with pytest.raises(ValueError, match="internal conversion bug"):
        reader.read_reviews(path)


def test_import_validates_entire_file_before_writing_any_raw_rows(tmp_path):
    """Persisting before the reader validates the complete file would create a partial invalid import."""
    from review_analytics.file_io.reader import read_reviews
    from review_analytics.repositories import ReviewRepository, initialize_database
    from review_analytics.services.ingestion import import_reviews
    from review_analytics.dto import ImportRequest

    path = tmp_path / "bad.csv"
    path.write_text("rating\n5\n", encoding="utf-8")
    database_path = tmp_path / "reviews.sqlite3"
    initialize_database(database_path)

    with pytest.raises(InputFileError):
        import_reviews(ImportRequest(path), ReviewRepository(database_path), read_reviews)

    with sqlite3.connect(database_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM raw_reviews").fetchone()[0] == 0


def test_blank_xlsx_review_cell_remains_missing_through_import_and_clean(tmp_path):
    """Converting an XLSX None cell to the literal text 'None' would corrupt raw meaning."""
    from review_analytics.dto import CleanRequest, ImportRequest
    from review_analytics.models import TargetMode
    from review_analytics.repositories import ReviewRepository, initialize_database
    from review_analytics.services.cleaning import clean_reviews
    from review_analytics.services.ingestion import import_reviews

    path = tmp_path / "blank-review.xlsx"
    workbook = Workbook()
    workbook.active.append(["review_text", "rating"])
    workbook.active.append([None, 3])
    workbook.save(path)
    database_path = tmp_path / "reviews.sqlite3"
    initialize_database(database_path)
    repository = ReviewRepository(database_path)

    imported = import_reviews(ImportRequest(path), repository)
    cleaned = clean_reviews(CleanRequest(TargetMode.PENDING), repository)

    assert imported.succeeded == 1
    assert cleaned.rejected == 1
    assert scalar(database_path, "SELECT review_text_raw FROM raw_reviews") == ""
    assert scalar(database_path, "SELECT rejection_reason FROM raw_reviews") == "MISSING_REVIEW_TEXT"


def test_xlsx_datetime_survives_raw_persistence_and_cleans_to_iso_date(tmp_path):
    """A valid Excel date must not become invalid after the raw SQLite round trip."""
    from review_analytics.dto import CleanRequest, ImportRequest
    from review_analytics.models import TargetMode
    from review_analytics.repositories import ReviewRepository, initialize_database
    from review_analytics.services.cleaning import clean_reviews
    from review_analytics.services.ingestion import import_reviews

    path = tmp_path / "dated-review.xlsx"
    workbook = Workbook()
    workbook.active.append(["review_text", "review_date"])
    workbook.active.append(["날짜가 있는 충분히 긴 리뷰", datetime(2026, 8, 1, 14, 30)])
    workbook.save(path)
    database_path = tmp_path / "reviews.sqlite3"
    initialize_database(database_path)
    repository = ReviewRepository(database_path)

    import_reviews(ImportRequest(path), repository)
    cleaned = clean_reviews(CleanRequest(TargetMode.PENDING), repository)

    assert cleaned.succeeded == 1
    assert cleaned.rejected == 0
    assert scalar(database_path, "SELECT review_date FROM clean_reviews") == "2026-08-01"


@pytest.mark.parametrize("export_format", (ExportFormat.CSV, ExportFormat.XLSX))
def test_write_export_creates_parent_and_headers_when_rows_are_empty(tmp_path, export_format):
    """An empty export must remain a usable file with its documented columns."""
    from review_analytics.file_io.exporter import write_export

    path = tmp_path / "nested" / f"reviews.{export_format.value}"

    generated = write_export((), export_format, path)

    assert generated.path == path
    assert generated.record_count == 0
    assert path.exists()
    if export_format is ExportFormat.CSV:
        assert path.read_bytes().startswith(b"\xef\xbb\xbf")
        assert path.read_text(encoding="utf-8-sig").splitlines() == [
            "review_id,review_text,rating,review_date,product_name,sentiment,confidence,analyzed_at"
        ]
    else:
        workbook = load_workbook(path, read_only=True)
        try:
            assert list(workbook.active.values) == [
                ("review_id", "review_text", "rating", "review_date", "product_name", "sentiment", "confidence", "analyzed_at")
            ]
        finally:
            workbook.close()


def test_write_export_serializes_flat_export_rows_to_utf8_bom_csv(tmp_path):
    """Writing enum objects or omitting fields would make the exported CSV unusable outside Python."""
    from review_analytics.file_io.exporter import write_export

    path = tmp_path / "reviews.csv"
    rows = (
        ExportRow(7, "좋아요", 5, "2026-08-01", "Bottle", Sentiment.POSITIVE, 0.95, "2026-08-02T00:00:00+00:00"),
    )

    generated = write_export(rows, ExportFormat.CSV, path)

    assert generated.record_count == 1
    assert path.read_bytes().startswith(b"\xef\xbb\xbf")
    assert path.read_text(encoding="utf-8-sig").splitlines()[1] == "7,좋아요,5,2026-08-01,Bottle,positive,0.95,2026-08-02T00:00:00+00:00"


def test_write_export_serializes_populated_rows_to_xlsx(tmp_path):
    """Writing only headers to XLSX would silently drop real export data."""
    from review_analytics.file_io.exporter import write_export

    path = tmp_path / "reviews.xlsx"
    rows = (ExportRow(8, "good bottle", 4, "2026-08-03", "Bottle", Sentiment.NEUTRAL, 0.5, None),)

    generated = write_export(rows, ExportFormat.XLSX, path)

    assert generated.record_count == 1
    workbook = load_workbook(path, read_only=True)
    try:
        assert list(workbook.active.values)[1] == (8, "good bottle", 4, "2026-08-03", "Bottle", "neutral", 0.5, None)
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("export_format", "suffix"),
    ((ExportFormat.CSV, ".xlsx"), (ExportFormat.XLSX, ".csv")),
)
def test_write_export_rejects_a_path_extension_that_conflicts_with_requested_format(tmp_path, export_format, suffix):
    """Writing one format under another extension makes a valid-looking export unreadable."""
    from review_analytics.file_io.exporter import write_export

    path = tmp_path / f"reviews{suffix}"

    with pytest.raises(OutputWriteError) as raised:
        write_export((), export_format, path)

    assert raised.value.code == "EXPORT_EXTENSION_MISMATCH"
    assert not path.exists()


def test_export_write_failure_logs_only_safe_output_metadata(tmp_path, caplog):
    from review_analytics.file_io.exporter import write_export

    blocker = tmp_path / "private-parent"
    blocker.write_text("PRIVATE FILE CONTENT", encoding="utf-8")
    output = blocker / "reviews.csv"

    with caplog.at_level(logging.ERROR), pytest.raises(OutputWriteError):
        write_export((), ExportFormat.CSV, output)

    assert "event=output.failed output_type=export file_name=reviews.csv error_code=EXPORT_WRITE_FAILED" in caplog.text
    assert "PRIVATE FILE CONTENT" not in caplog.text
    assert str(tmp_path) not in caplog.text

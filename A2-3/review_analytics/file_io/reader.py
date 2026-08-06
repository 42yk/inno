"""지원하는 리뷰 파일을 이름 있는 원본 도메인 입력으로 읽는다."""

from __future__ import annotations

import csv
import zipfile
from pathlib import Path
from xml.etree.ElementTree import ParseError

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from review_analytics.errors import InputFileError
from review_analytics.models import RawReviewInput


_OPTIONAL_COLUMNS = ("rating", "review_date", "product_name")
_REQUIRED_COLUMN = "review_text"


# 확장자에 맞는 리더로 리뷰 파일 전체를 원본 입력 모델로 읽는다.
def read_reviews(path: Path) -> tuple[RawReviewInput, ...]:
    """Fully validate *path* and return every input row without normalization."""
    input_path = Path(path)
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(input_path)
    if suffix == ".xlsx":
        return _read_xlsx(input_path)
    raise InputFileError("지원하지 않는 입력 파일 형식입니다.", "UNSUPPORTED_INPUT_FILE_TYPE")


# CSV 헤더와 각 행을 검증해 원본 리뷰 입력으로 변환한다.
def _read_csv(path: Path) -> tuple[RawReviewInput, ...]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream, strict=True)
            _validated_headers(reader.fieldnames)
            rows = tuple(
                RawReviewInput(
                    review_text_raw=row.get(_REQUIRED_COLUMN),
                    rating_raw=row.get("rating"),
                    review_date_raw=row.get("review_date"),
                    product_name_raw=row.get("product_name"),
                    source_type="csv",
                    source_ref=path.name,
                    source_row=row_number,
                )
                for row_number, row in enumerate(reader, start=2)
            )
            return rows
    except InputFileError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise InputFileError("입력 파일을 읽을 수 없습니다.", "INPUT_FILE_READ_FAILED") from exc


# XLSX 행을 헤더 위치에 맞춰 원본 리뷰 입력으로 변환한다.
def _read_xlsx(path: Path) -> tuple[RawReviewInput, ...]:
    values_by_row = _xlsx_values(path)
    headers = _validated_headers(values_by_row[0] if values_by_row else None)
    indexes = {name: headers.index(name) for name in (_REQUIRED_COLUMN, *_OPTIONAL_COLUMNS) if name in headers}
    return tuple(
        RawReviewInput(
            review_text_raw=_cell_value(values, indexes[_REQUIRED_COLUMN]),
            rating_raw=_cell_value(values, indexes.get("rating")),
            review_date_raw=_cell_value(values, indexes.get("review_date")),
            product_name_raw=_cell_value(values, indexes.get("product_name")),
            source_type="xlsx",
            source_ref=path.name,
            source_row=row_number,
        )
        for row_number, values in enumerate(values_by_row[1:], start=2)
    )


# XLSX 워크북을 안전하게 닫으며 셀 값 스냅샷을 읽는다.
def _xlsx_values(path: Path) -> tuple[tuple[object, ...], ...]:
    """Read workbook-owned row values while translating only parser failures."""
    try:
        with path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=False)
            try:
                return tuple(tuple(row) for row in workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
    except (
        OSError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
        ParseError,
        InvalidFileException,
    ) as exc:
        raise InputFileError("입력 파일을 읽을 수 없습니다.", "INPUT_FILE_READ_FAILED") from exc


# 필수 리뷰 본문 열이 있는지 확인하고 헤더 튜플을 반환한다.
def _validated_headers(headers: object) -> tuple[str, ...]:
    if headers is None:
        raise InputFileError("review_text 열이 필요합니다.", "MISSING_REVIEW_TEXT_COLUMN")
    values = tuple(headers)
    if _REQUIRED_COLUMN not in values:
        raise InputFileError("review_text 열이 필요합니다.", "MISSING_REVIEW_TEXT_COLUMN")
    return values


# 행 범위를 벗어난 선택 열을 None으로 처리해 셀 값을 얻는다.
def _cell_value(values: tuple[object, ...], index: int | None) -> object | None:
    return None if index is None or index >= len(values) else values[index]
